import os
import re
import warnings
import pandas as pd
import logging
from time import sleep, perf_counter
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries, get_column_letter
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from threading import Lock, Thread, Event
from sqlalchemy import create_engine
from sqlalchemy.exc import OperationalError
from APIs import historic_bars, company_facts, download_raw_ciks
from datetime import datetime
from Tools import print_progress, et_now
from Table_Tools import combine_all_tables, list_tables


# Base file path used for reading the source workbook and related assets.
f_path = r"C:\Portfolio Projects\Stock Analysis Tool"

# Establishing DB Path
db_path = 'company_facts.db'

@dataclass
class ProgressState:
    """Simple shared progress tracker for the threaded database build."""
    completed: int = 0
    total: int = 0

# Establishing Logging variables
logging.basicConfig(level=logging.INFO, format="[%(asctime)s] %(levelname)s:%(message)s")
db_logger = logging.getLogger(__name__)
formatter = logging.Formatter("[%(asctime)s] %(levelname)s:%(message)s")
file_handler = logging.FileHandler("DB_Output.log")
file_handler.setLevel(logging.DEBUG)
file_handler.setFormatter(formatter)

if not db_logger.handlers:
    db_logger.addHandler(file_handler)

db_logger.setLevel(logging.INFO)
db_logger.propagate = False

# Core label dictionaries used to translate SEC fact names into readable column names.
valueset = {'Total Assets': 'Assets',
            'Current Assets': 'AssetsCurrent',
            'Cash and Cash Equivalents': 'CashAndCashEquivalentsAtCarryingValue',
            'Accounts Receivable': 'AccountsReceivableNetCurrent',
            'Inventory': 'InventoryNet',
            'Total Non-Current Assets': 'NoncurrentAssets',
            'Property, Plant, and Equipment': 'PropertyPlantAndEquipmentNet',
            'Intangible Assets': 'IntangibleAssetsNetExcludingGoodwill',
            'Goodwill': 'Goodwill',
            'Total Liabilities': 'Liabilities',
            'Total Current Liabilities': 'LiabilitiesCurrent',
            'Current Accounts Payable': 'AccountsPayableCurrent',
            'Deferred Revenue': 'DeferredRevenue',
            'Current Debt': 'DebtCurrent',
            'Long Term Debt': 'LongTermDebt',
            'Total Shareholder Equity': 'StockholdersEquity',
            'Treasury Stock': 'TreasuryStockValue',
            'Retained Earnings': 'RetainedEarningsAccumulatedDeficit',
            'Common Stock': 'CommonStockValue',
            'Common Stock Shares Outstanding': 'CommonStockSharesOutstanding',
            'Gross Profit': 'GrossProfit',
            'Total Revenue': 'Revenues',
            'Cost of Revenue': 'CostOfRevenue',
            'Operating Income': 'OperatingIncomeLoss',
            'Research and Development': 'ResearchAndDevelopmentExpense',
            'Operating Expenses': 'OperatingExpenses',
            'Income Tax Expense': 'IncomeTaxExpenseBenefit',
            'Net Income': 'NetIncomeLoss',
            'Interest Expense': 'InterestExpense',
            'Depreciation & Amortization': 'DepreciationAndAmortization',
            'Operating Cash Flow': 'NetCashProvidedByUsedInOperatingActivities',
            'Change in Operating Liabilities': 'IncreaseDecreaseInAccruedLiabilitiesAndOtherOperatingLiabilities',
            'Depreciation, Depletion, and Amortization': 'DepreciationDepletionAndAmortization',
            'Capital Expenditures': 'PaymentsToAcquireProductiveAssets',
            'Cash Flow from Investment': 'NetCashProvidedByUsedInInvestingActivities',
            'Cash Flow from Financing': 'NetCashProvidedByUsedInFinancingActivities',
            'Dividend Payout': 'PaymentsOfDividends',
            'Dividend Payout (Common Stock)': 'DividendsCommonStockCash',
            'Earnings Per Share': 'EarningsPerShareBasic'
        }

# Subset of balance-sheet style values that can be copied into time-discrete quarter slots.
instant_valueset = {'Total Assets': 'Assets',
            'Current Assets': 'AssetsCurrent',
            'Cash and Cash Equivalents': 'CashAndCashEquivalentsAtCarryingValue',
            'Accounts Receivable': 'AccountsReceivableNetCurrent',
            'Inventory': 'InventoryNet',
            'Total Non-Current Assets': 'NoncurrentAssets',
            'Property, Plant, and Equipment': 'PropertyPlantAndEquipmentNet',
            'Intangible Assets': 'IntangibleAssetsNetExcludingGoodwill',
            'Goodwill': 'Goodwill',
            'Total Liabilities': 'Liabilities',
            'Total Current Liabilities': 'LiabilitiesCurrent',
            'Current Accounts Payable': 'AccountsPayableCurrent',
            'Deferred Revenue': 'DeferredRevenue',
            'Current Debt': 'DebtCurrent',
            'Long Term Debt': 'LongTermDebt',
            'Total Shareholder Equity': 'StockholdersEquity',
            'Treasury Stock': 'TreasuryStockValue',
            'Retained Earnings': 'RetainedEarningsAccumulatedDeficit'
        }

# Helper utilities for safe value handling and gap-filling.
def get_fy(years_map, year_int):
    """Return a flag and the FY value for a requested year from a nested year map."""
    logging.debug(f"Checking FY availability for year {year_int}.")
    if str(year_int) not in years_map:
        db_logger.debug(f"No FY entry found for year {year_int}.")
        return 0, None

    fy_value = years_map.get(str(year_int), {}).get("FY")
    logging.debug(f"FY lookup for year {year_int} returned {fy_value}.")
    return 1, fy_value

def nan_to_none(v):
    """Convert pandas missing values into Python None for safer downstream handling."""
    result = None if pd.isna(v) else v
    logging.debug(f"nan_to_none converted {v} to {result}.")
    return result

def safe_div(numerator, denominator):
    """Safely divide two values and return None when the operation is not valid."""
    logging.debug(f"safe_div received numerator={numerator}, denominator={denominator}.")
    if pd.isna(numerator) or pd.isna(denominator) or float(denominator) == 0:
        db_logger.debug("safe_div returned None because numerator/denominator was invalid.")
        return None
    float_num = float(numerator)
    float_den = float(denominator)
    result = float_num / float_den
    logging.debug(f"safe_div returning result {result}.")
    return result

def fill_nulls(df, unique_tickers, fillable_values):
    """Fill missing annual values by interpolating across nearby fiscal-year observations."""
    total_tickers = len(unique_tickers)
    start_time = perf_counter()
    db_logger.info(f"Starting fill_nulls for {total_tickers} tickers.")
    logging.debug("Entering fill_nulls.")

    for idx, ticker in enumerate(unique_tickers, start=1):
        logging.debug(f"Processing null fills for ticker {ticker} ({idx}/{total_tickers}).")
        try:
            ticker_df = df[df['Ticker'] == ticker].reset_index(drop=True)
            value_dict = {value: {} for value in fillable_values}
            for value in fillable_values:
                for _, row in ticker_df.iterrows():
                    year = row["Year"]
                    period = row["Period"]
                    amount = nan_to_none(row[value])
                    value_dict[value].setdefault(year, {})
                    value_dict[value][year][period] = amount if amount is not None else None

                for primary_label, years_map in value_dict.items():
                    logging.debug(f"Evaluating null-fill label {primary_label} for ticker {ticker}.")
                    for year, periods_map in years_map.items():
                        for variable in ["Q1", "Q2", "Q3", "FY"]:
                            if variable not in periods_map:
                                periods_map[variable] = None

                        FY = float(periods_map["FY"]) if periods_map["FY"] is not None else None
                        yi = int(year)

                        if FY is None:
                            back_value = None
                            forward_value = None
                            count_b = -1
                            while True:
                                found, back_value = get_fy(years_map, yi + count_b)
                                if found == 0:
                                    break
                                if back_value is not None:
                                    break
                                count_b -= 1

                            count_f = 1
                            while True:
                                found, forward_value = get_fy(years_map, yi + count_f)
                                if found == 0:
                                    break
                                if forward_value is not None:
                                    break
                                count_f += 1

                            if back_value is not None and forward_value is not None:
                                interpolation_range = count_f + abs(count_b)
                                position = abs(count_b)
                                modifier = (forward_value - back_value) * (position / interpolation_range)
                                periods_map["FY"] = back_value + modifier
                                logging.debug(f"Interpolated FY value for ticker {ticker}, year {year}, label {primary_label}.")
                            else:
                                periods_map["FY"] = None
                                db_logger.debug(f"Unable to interpolate FY value for ticker {ticker}, year {year}, label {primary_label}.")

                        for period in ["FY"]:
                            new_value = periods_map[period]
                            match = (
                                (df["Ticker"] == ticker) &
                                (df["Year"].astype(str) == str(year)) &
                                (df["Period"] == period) &
                                (df[value].isna())
                            )
                            df.loc[match, value] = new_value

            print_progress('Filling Nulls', idx, total_tickers, start_time)
        except Exception:
            db_logger.error(f"Failed while filling nulls for ticker {ticker}.")
            raise

    db_logger.info("Completed fill_nulls.")
    logging.debug("Leaving fill_nulls.")
    return df

# Source-loading functions that gather company metadata and market-price history.
def get_company_list():
    """Read the source Excel table and merge it with downloaded CIK reference data."""
    db_logger.info("Loading company list from workbook source.")
    logging.debug("Entering get_company_list.")

    try:
        # Define Variables
        path = os.path.join(f_path, "Company Analysis.xlsm")
        sheet_name = "Source"
        table_name = "STable"
        logging.debug(f"Workbook path resolved to {path}.")

        # Filter out formatting warnings
        warnings.filterwarnings(
            "ignore",
            category=UserWarning,
            module="openpyxl.worksheet._reader",
            message=".*extension is not supported.*"
        )

        # Read Workbook metadata
        wb = load_workbook(path, read_only=False, data_only=True)
        ws = wb[sheet_name]
        logging.debug("Workbook loaded successfully.")

        # Get Table Range
        table_ref = ws.tables[table_name].ref
        min_col, min_row, max_col, max_row = range_boundaries(table_ref)
        usecols = f"{get_column_letter(min_col)}:{get_column_letter(max_col)}"
        skiprows = min_row - 1
        nrows = max_row - min_row + 1

        excel_df = pd.read_excel(
            path,
            sheet_name=sheet_name,
            usecols=usecols,
            skiprows=skiprows,
            nrows=nrows,
            header=0,
            engine="openpyxl"
        )
        logging.debug(f"Loaded {len(excel_df)} source rows from Excel.")

        ciks = download_raw_ciks()
        ciks_df = pd.DataFrame(ciks)[["Symbol", "CIK"]]
        logging.debug(f"Downloaded {len(ciks_df)} raw CIK records.")

        combined_df = excel_df.merge(ciks_df.rename(columns={"Symbol": "Ticker"}))[["Ticker", "CIK", "Sector"]]

        db_logger.info(f"Prepared combined company list with {len(combined_df)} rows.")
        return combined_df
    except Exception:
        db_logger.error("Failed to load company list.")
        raise

def get_shareprices(ticker, starting_year):
    """Fetch yearly and quarterly VWAP share-price history for a single ticker."""
    start = f"{starting_year}-01-01"
    end = et_now().strftime("%Y-%m-%d")

    db_logger.info(f"Fetching share price history for {ticker} starting {start}.")
    logging.debug(f"Entering get_shareprices for ticker {ticker}.")

    try:
        q_price_data = historic_bars(
            symbol=ticker,
            start=start,
            end=end,
            timeframe='3Month'
        )

        y_price_data = historic_bars(
            symbol=ticker,
            start=start,
            end=end,
            timeframe='12Month'
        )

        q_bars = q_price_data.get("bars", [])
        y_bars = y_price_data.get("bars", [])

        if isinstance(q_bars, dict):
            q_price_list = q_bars.get(ticker, [])
        else:
            q_price_list = q_bars

        if isinstance(y_bars, dict):
            y_price_list = y_bars.get(ticker, [])
        else:
            y_price_list = y_bars

        price_dict = {"Year": {}, "Quarters": {}}

        for section in y_price_list:
            year = section["t"].split("-")[0]
            price_dict["Year"][year] = section["vw"]

        for section in q_price_list:
            year = section["t"].split("-")[0]
            month = int(section["t"].split("-")[1])

            if month > 1 and month <= 4:
                quarter = "Q1"
            elif month > 4 and month <= 7:
                quarter = "Q2"
            elif month > 7 and month <= 10:
                quarter = "Q3"
            else:
                quarter = "Q4"
            price_dict["Quarters"][f"{year}-{quarter}"] = section["vw"]

        db_logger.info(f"Retrieved {len(y_price_list)} annual bars and {len(q_price_list)} quarterly bars for {ticker}.")
        logging.debug(f"Leaving get_shareprices for ticker {ticker}.")
        return price_dict
    except Exception:
        db_logger.error(f"Failed to retrieve share prices for ticker {ticker}.")
        raise


# Core transformation functions that build financial rows and calculate ratios.
def value_pulls(cik, label_data, comp_data):
    """
    Build a period-by-period financial dataset for one company from SEC facts data.

    The function initializes yearly and quarterly rows, maps raw fact labels into
    readable fields, attaches share prices, and computes helper fields such as
    average inventory for later ratio calculations.
    """
    db_logger.info(f"Starting value pull for CIK {cik}.")
    logging.debug(f"Entering value_pulls for CIK {cik}.")
    value_dict = {}

    current_year = 2026
    starting_year = current_year - 10

    if label_data == {}:
        db_logger.warning(f"No label data returned for CIK {cik}.")
        logging.debug("value_pulls exited early because label_data was empty.")
        return None, None

    try:
        ticker = comp_data['Ticker']
        price_dict = get_shareprices(ticker, starting_year)
        logging.debug(f"Share price dictionary built for ticker {ticker}.")

        for year in range(current_year, starting_year, -1):
            logging.debug(f"Initializing value slots for CIK {cik}, year {year}.")
            for period in ["FY", "Q4-TD", "Q4-SQ", "Q3-SQ", "Q3-TD", "Q3-SQ", "Q2-TD", "Q2-SQ", "Q1-SQ"]:
                primary_key = f"{cik}-{year}-{period}"
                value_dict[primary_key] = {}
                value_dict[primary_key]['CIK'] = cik
                value_dict[primary_key]['Name'] = comp_data['Company']
                value_dict[primary_key]['Ticker'] = comp_data['Ticker']
                value_dict[primary_key]['Sector'] = comp_data['Sector']
                value_dict[primary_key]['Industry'] = comp_data['Industry']
                value_dict[primary_key]['SIC'] = comp_data['SIC Code']
                value_dict[primary_key]['Address'] = comp_data['Address']
                value_dict[primary_key]['Year'] = year
                value_dict[primary_key]['Period'] = period

                if period == "FY":
                    value_dict[primary_key]['Share Price'] = price_dict["Year"].get(str(year))
                else:
                    fixed_period = f"{year}-{period.split('-')[0]}"
                    value_dict[primary_key]['Share Price'] = price_dict["Quarters"].get(fixed_period)

                for label in valueset.keys():
                    value_dict[primary_key][label] = None

        for label, title in valueset.items():
            logging.debug(f"Processing label mapping {label} -> {title} for CIK {cik}.")
            if title not in label_data:
                db_logger.debug(f"Missing source tag {title} for CIK {cik}.")
                continue
            else:
                if 'USD' in label_data[title]['units']:
                    bracket = label_data[title]['units']['USD']
                elif 'USD/shares' in label_data[title]['units']:
                    bracket = label_data[title]['units']['USD/shares']
                elif 'shares' in label_data[title]['units']:
                    bracket = label_data[title]['units']['shares']
                else:
                    db_logger.debug(f"No supported unit bucket found for {title} in CIK {cik}.")
                    continue

                for line in bracket:
                    form = line["form"]
                    if form != '10-K' and form != '10-Q':
                        continue
                    else:
                        year = line["fy"]
                        if int(year) < current_year - 9:
                            continue
                        value = line["val"]
                        period = line["fp"]
                        if period == "FY":
                            value_dict[f"{cik}-{year}-FY"][label] = value
                            continue
                        start = line["start"] if "start" in line else None
                        end = line["end"]
                        period_attach = "-SQ"
                        if start is not None and period != 'Q1':
                            start_date = datetime.strptime(start, '%Y-%m-%d')
                            end_date = datetime.strptime(end, '%Y-%m-%d')
                            if (end_date - start_date).days > 35:
                                period_attach = "-TD"

                        full_period = period + period_attach
                        value_dict[f"{cik}-{year}-{full_period}"][label] = value

                        if label in instant_valueset and period in ["Q2", "Q3", "Q4"] and period_attach == "-SQ":
                            if value_dict[f"{cik}-{year}-{period}-TD"][label] is None:
                                value_dict[f"{cik}-{year}-{period}-TD"][label] = value

        logging.debug(f"Starting derived field calculations for CIK {cik}.")
        for row_key in value_dict.keys():
            average_inventory = None
            year = value_dict[row_key]["Year"]
            current_inventory = value_dict[row_key]["Inventory"]
            previous_year = int(year) - 1

            prev_key = f"{cik}-{previous_year}-FY"
            if prev_key in value_dict:
                previous_inventory = value_dict[prev_key]["Inventory"]
                if current_inventory is not None and previous_inventory is not None:
                    average_inventory = (current_inventory + previous_inventory) / 2

            value_dict[row_key]["Average Inventory"] = average_inventory

        df = pd.DataFrame.from_dict(value_dict, orient='index')
        db_logger.info(f"Completed value pull for CIK {cik} with {len(df)} rows.")
        logging.debug(f"Leaving value_pulls for CIK {cik}.")
        return value_dict, df
    except Exception:
        db_logger.error(f"value_pulls failed for CIK {cik}.")
        raise

def formula_calculator(df):
    """Calculate operating, investing, financing, and valuation ratios for the dataset."""
    db_logger.info("Starting formula calculations.")
    logging.debug("Entering formula_calculator.")

    if df is None or df.empty:
        db_logger.warning("formula_calculator received an empty dataframe.")
        return df

    try:
        logging.debug("Applying operating formulas.")
        df['Operating Margin'] = df.apply(
            lambda row: safe_div(row['Operating Income'], row['Total Revenue']),
            axis=1
        )

        df['EBIT'] = df.apply(
            lambda row: (row['Net Income'] + row['Interest Expense'] + row['Income Tax Expense'])
            if pd.notna(row['Net Income']) and pd.notna(row['Interest Expense']) and pd.notna(row['Income Tax Expense'])
            else None,
            axis=1
        )

        df['EBITDA'] = df.apply(
            lambda row: (row['Net Income'] + row['Interest Expense'] + row['Income Tax Expense'] + row['Depreciation & Amortization'])
            if pd.notna(row['Net Income']) and pd.notna(row['Interest Expense']) and pd.notna(row['Income Tax Expense']) and pd.notna(row['Depreciation & Amortization'])
            else row['EBIT'],
            axis=1
        )

        df['EBITDA Margin'] = df.apply(
            lambda row: safe_div(row['EBITDA'], row['Total Revenue']),
            axis=1
        )

        df['Operating CF Margin'] = df.apply(
            lambda row: safe_div(row['Operating Cash Flow'], row['Total Revenue']),
            axis=1
        )

        df['Free CF Margin'] = df.apply(
            lambda row: safe_div(
                (row['Operating Cash Flow'] - row['Capital Expenditures'])
                if pd.notna(row['Operating Cash Flow']) and pd.notna(row['Capital Expenditures'])
                else None,
                row['Total Revenue']
            ),
            axis=1
        )

        df['Gross Margin'] = df.apply(
            lambda row: safe_div(row['Gross Profit'], row['Total Revenue']),
            axis=1
        )

        df['Inventory Turnover'] = df.apply(
            lambda row: safe_div(row['Cost of Revenue'], row['Average Inventory']),
            axis=1
        )

        df['Asset Turnover'] = df.apply(
            lambda row: safe_div(row['Total Revenue'], row['Total Assets']),
            axis=1
        )

        logging.debug("Applying investing formulas.")
        df['Investing CF Ratio'] = df.apply(
            lambda row: safe_div(row['Cash Flow from Investment'], row['Total Revenue']),
            axis=1
        )

        df['CapEx Ratio'] = df.apply(
            lambda row: safe_div(row['Capital Expenditures'], row['Total Revenue']),
            axis=1
        )

        df['CapEx Coverage Ratio'] = df.apply(
            lambda row: safe_div(row['Operating Cash Flow'], row['Capital Expenditures']),
            axis=1
        )

        df['R&D Margin'] = df.apply(
            lambda row: safe_div(row['Research and Development'], row['Total Revenue']),
            axis=1
        )

        logging.debug("Applying financing formulas.")
        df['Debt Ratio'] = df.apply(
            lambda row: safe_div(row['Total Liabilities'], row['Total Assets']),
            axis=1
        )

        df['Interest Coverage Ratio'] = df.apply(
            lambda row: safe_div(row['EBIT'], row['Interest Expense']),
            axis=1
        )

        df['Cash Ratio'] = df.apply(
            lambda row: safe_div(row['Cash and Cash Equivalents'], row['Total Current Liabilities']),
            axis=1
        )

        df['Current Ratio'] = df.apply(
            lambda row: safe_div(
                (row['Current Assets'] - row['Inventory'])
                if pd.notna(row['Current Assets']) and pd.notna(row['Inventory'])
                else None,
                row['Total Current Liabilities']
            ),
            axis=1
        )

        logging.debug("Applying shareholder relation formulas.")
        df['Price to Earnings Ratio'] = df.apply(
            lambda row: safe_div(row['Share Price'], row['Earnings Per Share']),
            axis=1
        )

        df['Return on Equity'] = df.apply(
            lambda row: safe_div(row['Net Income'], row['Total Shareholder Equity']),
            axis=1
        )

        df['Dividend Yield'] = df.apply(
            lambda row: safe_div(row['Dividend Payout'], row['Share Price']),
            axis=1
        )

        df['Price to Book Ratio'] = df.apply(
            lambda row: safe_div(
                row['Share Price'],
                safe_div(row['Total Shareholder Equity'], row['Earnings Per Share'])
            ),
            axis=1
        )

        db_logger.info(f"Completed formula calculations for {len(df)} rows.")
        logging.debug("Leaving formula_calculator.")
        return df
    except Exception:
        db_logger.error("formula_calculator failed.")
        raise

# Persistence and orchestration helpers for writing sector results into SQLite.
def normalize_table_name(name):
    """Convert a sector name into a SQLite-safe table name."""
    logging.debug(f"Normalizing SQL table name for sector {name}.")
    table = re.sub(r"[^0-9a-zA-Z_]+", "_", str(name).strip().lower())
    if not table:
        table = "unknown_sector"
    if table[0].isdigit():
        table = f"sector_{table}"
    db_logger.debug(f"Normalized table name '{name}' to '{table}'.")
    return table

def write_table_with_retry(df, table_name, engine, write_lock, attempts=5, base_delay=0.5):
    """Write a dataframe to SQLite and retry if the database is temporarily locked."""
    db_logger.info(f"Writing dataframe to SQL table {table_name}.")
    for attempt in range(1, attempts + 1):
        try:
            logging.debug(f"Attempt {attempt} writing {len(df)} rows to {table_name}.")
            with write_lock:
                df.to_sql(table_name, con=engine, if_exists="replace", index=False)
            db_logger.info(f"Successfully wrote table {table_name} on attempt {attempt}.")
            return
        except OperationalError as e:
            msg = str(e).lower()
            db_logger.warning(f"Write attempt {attempt} failed for {table_name}: {e}")
            if "locked" not in msg or attempt == attempts:
                db_logger.error(f"Unable to write SQL table {table_name} after {attempt} attempts.")
                raise
            sleep(base_delay * attempt)

def process_sector(engine, write_lock, sector, sector_df, progress_state, progress_lock):
    """Process every company in a sector, calculate results, and persist them to SQL."""
    db_logger.info(f"Starting sector processing for {sector}.")
    logging.debug(f"Entering process_sector for {sector} with {len(sector_df)} companies.")
    cik_list = sector_df["CIK"]
    full_df = None

    for cik in cik_list:
        comp_data = None
        try:
            logging.debug(f"Fetching company facts for sector {sector}, CIK {cik}.")
            comp_data, label_data = company_facts(cik)
            value_dict, df = value_pulls(cik, label_data, comp_data)
            if value_dict is None:
                db_logger.warning(f"No value data returned for sector {sector}, CIK {cik}.")
                continue

            formula_df = formula_calculator(df)
            if full_df is None:
                full_df = formula_df
            else:
                full_df = pd.concat([full_df, formula_df], ignore_index=False)
            db_logger.debug(f"Appended processed dataframe for sector {sector}, CIK {cik}.")

        except Exception as E:
            company_name = comp_data['Company'] if isinstance(comp_data, dict) and 'Company' in comp_data else 'UNKNOWN'
            db_logger.error(f"Error processing sector {sector}, CIK {cik}, company {company_name}: {E}")

        finally:
            with progress_lock:
                progress_state.completed += 1

    if full_df is not None and not full_df.empty:
        table_name = normalize_table_name(sector)
        write_table_with_retry(full_df, table_name, engine, write_lock)
        db_logger.info(f"Finished sector processing for {sector} with {len(full_df)} rows written.")
    else:
        db_logger.warning(f"Sector {sector} produced no output rows.")

def progress_printer(progress_state, progress_lock, start_time, done_event):
    """Continuously print build progress until all worker threads finish."""
    db_logger.debug("Progress printer started.")
    while not done_event.is_set():
        with progress_lock:
            completed = progress_state.completed
            total = progress_state.total
        logging.debug(f"Database build progress: {completed}/{total}.")
        print_progress("Building Databases", completed, total, start_time)
        sleep(0.2)

    with progress_lock:
        completed = progress_state.completed
        total = progress_state.total
    print_progress("Building Databases", completed, total, start_time)
    db_logger.debug("Progress printer finished.")


# Main workflow: gather the company universe, split work by sector, and build the database.
def main():
    """Coordinate the full multi-threaded database build process from start to finish."""
    db_logger.info("Starting database build workflow.")
    logging.debug("Entering main.")
    company_df = get_company_list()
    sector_list = company_df["Sector"].dropna().unique().tolist()
    db_logger.info(f"Loaded {len(company_df)} companies across {len(sector_list)} sectors.")

    engine = create_engine(
        "sqlite:///company_facts.db",
        connect_args={"timeout": 60, "check_same_thread": False},
        pool_pre_ping=True
    )
    write_lock = Lock()

    progress_state = ProgressState(total=len(company_df["CIK"]))
    progress_lock = Lock()
    done_event = Event()
    start_time = perf_counter()

    printer = Thread(
        target=progress_printer,
        args=(progress_state, progress_lock, start_time, done_event),
        daemon=True
    )
    printer.start()

    try:
        with ThreadPoolExecutor(max_workers=min(3, max(1, len(sector_list)))) as pool:
            futures = []
            for sector in sector_list:
                logging.debug(f"Submitting sector {sector} to thread pool.")
                sector_df = company_df[company_df["Sector"] == sector].reset_index(drop=True)
                futures.append(
                    pool.submit(process_sector, engine, write_lock, sector, sector_df, progress_state, progress_lock)
                )

            for f in as_completed(futures):
                f.result()

        db_logger.info("Database build workflow completed successfully.")
    except Exception:
        db_logger.error("Database build workflow failed.")
        raise
    finally:
        done_event.set()
        printer.join()
        combine_all_tables(db_path)
        list_tables(db_path)
        logging.debug("Leaving main.")


if __name__ == "__main__":
    main()